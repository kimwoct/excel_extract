import pandas as pd
import re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

def process_timetable(input_file, output_file):
    """
    Process a teacher timetable Excel file and convert it to a normalized format
    where each timeslot has exactly one row.
    
    Args:
        input_file: Path to the input Excel file
        output_file: Path to save the normalized Excel file
    """
    # Read the Excel file
    xls = pd.ExcelFile(input_file)
    
    # Create a workbook for output
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet
    
    # Process each sheet (each teacher's schedule)
    for sheet_name in xls.sheet_names:
        # Read the sheet
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
        
        # Find where the actual table begins (looking for "課節" or similar)
        start_row = None
        for i in range(min(20, len(df))):
            if isinstance(df.iloc[i, 0], str) and df.iloc[i, 0] == "課節":
                start_row = i
                break
        
        if start_row is None:
            print(f"Could not find start of table in sheet {sheet_name}, skipping")
            continue
        
        # Get the teacher name from cells above the table
        teacher_name = None
        for i in range(start_row):
            if isinstance(df.iloc[i, 0], str) and "老師" in df.iloc[i, 0]:
                teacher_name = df.iloc[i, 0]
                break
        
        if teacher_name is None:
            teacher_name = sheet_name
        
        # Extract column headers (days of week)
        headers = [str(df.iloc[start_row, j]) if pd.notna(df.iloc[start_row, j]) else "" 
                  for j in range(df.shape[1])]
        
        # Create a new sheet for this teacher
        ws = wb.create_sheet(title=sheet_name)
        
        # Write the teacher name and headers
        ws.cell(row=1, column=1, value=teacher_name)
        for j, header in enumerate(headers):
            ws.cell(row=2, column=j+1, value=header)
        
        # Define style for headers
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Process and normalize the schedule data
        normalized_data = []
        current_row = start_row + 1
        output_row = 3  # Start writing data from row 3
        
        while current_row < len(df):
            # Check if this is a timeslot row
            time_pattern = r'^\d{2}:\d{2}-\d{2}:\d{2}$'
            if pd.notna(df.iloc[current_row, 1]) and isinstance(df.iloc[current_row, 1], str) and re.match(time_pattern, df.iloc[current_row, 1]):
                # This is a timeslot row
                timeslot = df.iloc[current_row, 1]
                period = df.iloc[current_row, 0] if pd.notna(df.iloc[current_row, 0]) else ""
                
                # Get all content for this timeslot
                timeslot_data = [period, timeslot]
                for col in range(2, df.shape[1]):
                    # Collect all non-empty cell values from this column for current timeslot
                    values = []
                    row_offset = 0
                    while (current_row + row_offset < len(df) and 
                           (row_offset == 0 or 
                            not pd.notna(df.iloc[current_row + row_offset, 1]) or 
                            not isinstance(df.iloc[current_row + row_offset, 1], str) or 
                            not re.match(time_pattern, df.iloc[current_row + row_offset, 1]))):
                        if pd.notna(df.iloc[current_row + row_offset, col]):
                            value = str(df.iloc[current_row + row_offset, col])
                            if value.strip():  # Only add non-empty strings
                                values.append(value)
                        row_offset += 1
                        if current_row + row_offset >= len(df):
                            break
                    
                    # Join all values with newlines
                    timeslot_data.append("\n".join(values))
                
                # Write data to worksheet
                for j, value in enumerate(timeslot_data):
                    cell = ws.cell(row=output_row, column=j+1, value=value)
                    if j > 1:  # For day columns, enable text wrapping
                        cell.alignment = Alignment(wrap_text=True, vertical='center')
                
                # Move to next output row
                output_row += 1
                
                # Skip any rows that were part of this timeslot
                row_offset = 1
                while (current_row + row_offset < len(df) and 
                       (not pd.notna(df.iloc[current_row + row_offset, 1]) or 
                        not isinstance(df.iloc[current_row + row_offset, 1], str) or 
                        not re.match(time_pattern, df.iloc[current_row + row_offset, 1]))):
                    row_offset += 1
                
                current_row += row_offset
            else:
                current_row += 1
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    cell_length = len(str(cell.value).split('\n')[0])
                    max_length = max(max_length, min(cell_length, 30))
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        
        # Set appropriate row heights
        for row in ws.iter_rows(min_row=3, max_row=output_row):
            max_lines = 1
            for cell in row:
                if cell.value:
                    lines = str(cell.value).count('\n') + 1
                    max_lines = max(max_lines, lines)
            ws.row_dimensions[cell.row].height = max(15 * max_lines, 20)
    
    # Save the workbook
    wb.save(output_file)
    print(f"Transformation complete. Output saved to {output_file}")

# Example usage
if __name__ == "__main__":
    input_file = "04-Ge-Jiao-Shi-Shou-Ke-Shi-Jian-Biao.xlsx"
    output_file = "transformed_timetable.xlsx"
    process_timetable(input_file, output_file)